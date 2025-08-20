import streamlit as st 
import pandas as pd
import os
import time
from datetime import datetime
from openpyxl import load_workbook
# =======================
# USER LOGIN SYSTEM
# =======================
USER_CREDENTIALS = {
    "admin": "Ashutosh",
    "Ashutosh": "police@2025",
    "officer": "medal123"
}

def login():
    st.title("🔐 Police Medalist Data - Login")

    username = st.text_input("👤 Username")
    password = st.text_input("🔑 Password", type="password")

    if st.button("Login"):
        if username in USER_CREDENTIALS and USER_CREDENTIALS[username] == password:
            st.session_state["logged_in"] = True
            st.session_state["username"] = username
            st.success("✅ Login successful!")
            st.rerun()
        else:
            st.error("❌ Invalid Username or Password")

# =======================
# MAIN APP WITH LOGIN
# =======================
if "logged_in" not in st.session_state or not st.session_state["logged_in"]:
    login()
else:
    st.sidebar.success(f"👋 Welcome, {st.session_state['username']}")

    if st.sidebar.button("🚪 Logout"):
        st.session_state.clear()
        st.rerun()

    # ======== Place your existing app code here ========
    st.title("🏅 Uttarakhand Police Medalist Data Search and Update")
    st.write("Now your secured app starts here...")

# =======================
# File name
# =======================
FILE_NAME = "Main_Data.xlsx"

# =======================
# Safe save function (preserve formatting)
# =======================
def save_data_preserve_format(df, file_name=FILE_NAME):
    while True:
        try:
            if not os.path.exists(file_name):
                # First-time creation with pandas (no formatting to preserve yet)
                df.to_excel(file_name, index=False)
                break

            wb = load_workbook(file_name)
            ws = wb.active

            # Write headers (row 1)
            for c_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=1, column=c_idx, value=col_name)

            # Clear old values (rows except header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None

            # Write DataFrame values
            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx+2, column=c_idx, value=value)  # +2 = after header row

            wb.save(file_name)
            break

        except PermissionError:
            time.sleep(1)

# =======================
# Date format validation
# =======================
def validate_date(date_str):
    try:
        datetime.strptime(date_str, "%d-%m-%Y")
        return True
    except ValueError:
        return False

# =======================
# Auto-reload mechanism
# =======================
def get_file_timestamp():
    return os.path.getmtime(FILE_NAME) if os.path.exists(FILE_NAME) else 0

@st.cache_data
def load_data(timestamp):
    if os.path.exists(FILE_NAME):
        return pd.read_excel(FILE_NAME)
    else:
        return pd.DataFrame(columns=[
            "S.N", "नाम अधिकारी", "पदनाम", "पदनाम एवं नियुक्ति",
            "पदक का आधार", "पदक का नाम", "पदक का वर्ष",
            "जन्मतिथि", "भर्ती तिथि"
        ])

# =======================
# Auto assign serial numbers
# =======================
def assign_serial_numbers(df):
    df = df.reset_index(drop=True)
    df["S.N"] = range(1, len(df) + 1)
    return df

# =======================
# Streamlit App
# =======================
st.set_page_config(page_title="Police Medal Search", layout="wide")
st.title("🏅 Uttarakhand Police Medalist Data Search and Update")

# Load data with timestamp check
df = load_data(get_file_timestamp())
df = assign_serial_numbers(df)

# =======================
# Show Results
# =======================
st.subheader("📋 Medalist Records")

# Inject CSS for wrap text in data editor
st.markdown("""
    <style>
    [data-testid="stDataFrame"] div {
        white-space: normal !important;
        line-height: 1.4em !important;
        word-wrap: break-word !important;
    }
    </style>
    """, unsafe_allow_html=True)

with st.container():
    # Search bar always visible
    cols = st.columns([3,1])
    with cols[0]:
        search_term = st.text_input(
            "🔍 Enter name, designation, Date of Birth or Date of Joining", 
            key="search_box"
        ).strip()

    if search_term:
        mask = df.apply(lambda row: row.astype(str).str.contains(search_term, case=False, na=False)).any(axis=1)
        search_results = df[mask]
    else:
        search_results = df.copy()

    if search_results.empty:
        st.warning("⚠️ No Data Found. Please enter the correct keyword in Hindi.")
        edited_df = pd.DataFrame(columns=df.columns)
    else:
        enable_edit = st.checkbox("✏️ Enable Editing")

        # Columns that should NEVER be editable
        always_disabled = ["S.N", "पदक का आधार", "पदक का नाम", "पदक का वर्ष"]
        disabled_cols = always_disabled if enable_edit else list(search_results.columns)

        # Medalist Records Table
        edited_df = st.data_editor(
            search_results,
            num_rows="dynamic",
            use_container_width=True,
            disabled=disabled_cols
        )

        # Detect changes
                # Detect changes
        if not edited_df.equals(search_results):
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Save Changes", type="primary"):

                    # ===== Date Validation before Saving =====
                    invalid_dates = []
                    for idx, row in edited_df.iterrows():
                        for col in ["जन्मतिथि", "भर्ती तिथि"]:
                            if pd.notna(row[col]) and str(row[col]).strip() != "":
                                if not validate_date(str(row[col])):
                                    invalid_dates.append(f"{col} → {row[col]}")

                    if invalid_dates:
                        st.error("❌ कृपया जन्मतिथि एवं भर्ती तिथि को सही DD-MM-YYYY फार्मेट में भरे")
                        st.write("⚠️ गलत डेटा:", invalid_dates)
                    else:
                        if search_term:
                            df.loc[mask] = edited_df.values
                        else:
                            df = edited_df

                        df = assign_serial_numbers(df)
                        save_data_preserve_format(df)
                        st.success("✅ Data saved successfully!")
                        st.cache_data.clear()
                        st.rerun()

            with col2:
                if st.button("❌ Discard Changes"):
                    st.cache_data.clear()
                    st.rerun()

# =======================
# Add New Row
# =======================
st.subheader("➕ Add New Entry for Medal Data!")

with st.form("new_entry_form"):
    new_data = {}

    with st.expander("👤 Personal Details"):
        new_data["नाम अधिकारी"] = st.text_input("नाम अधिकारी", key="new_name")
        new_data["जन्मतिथि"] = st.text_input("जन्मतिथि (DD-MM-YYYY)", key="new_dob")
        new_data["भर्ती तिथि"] = st.text_input("भर्ती तिथि (DD-MM-YYYY)", key="new_doj")

    with st.expander("🏅 Medal Details"):
        new_data["पदनाम"] = st.text_input("पदनाम", key="new_designation")
        new_data["पदनाम एवं नियुक्ति"] = st.text_input("पदनाम एवं नियुक्ति", key="new_posting")
        new_data["पदक का आधार"] = st.text_input("पदक का आधार", key="new_medal_basis")
        new_data["पदक का नाम"] = st.text_input("पदक का नाम", key="new_medal_name")
        new_data["पदक का वर्ष"] = st.text_input("पदक का वर्ष", key="new_medal_year")

    submitted = st.form_submit_button("Add Row")

    if submitted:
        invalid = False
        # Date validation
        if new_data["जन्मतिथि"] and not validate_date(new_data["जन्मतिथि"]):
            st.error("❌ Invalid जन्मतिथि! Please use DD-MM-YYYY")
            invalid = True
        if new_data["भर्ती तिथि"] and not validate_date(new_data["भर्ती तिथि"]):
            st.error("❌ Invalid भर्ती तिथि! Please use DD-MM-YYYY")
            invalid = True

        if not invalid:
            # Append new row
            df = pd.concat([df, pd.DataFrame([new_data], columns=df.columns)], ignore_index=True)
            df = assign_serial_numbers(df)
            save_data_preserve_format(df)
            st.success("✅ New row added successfully!")
            st.cache_data.clear()
            st.rerun()

# =======================
# Medal Data Summary
# =======================
# =======================
# Medal Data Summary
# =======================
st.subheader("📊 Medal Distribution Data")

if st.button("🏅 पदक डाटा टेबल"):
    if df.empty:
        st.warning("⚠️ कोई डेटा उपलब्ध नहीं है।")
    else:
        # Group by Medal Name + Year
        medal_summary = (
            df.groupby(["पदक का वर्ष", "पदक का नाम"])
              .size()
              .reset_index(name="पदक की संख्या")
              .sort_values("पदक का वर्ष")
	  ) 
# Year filter
        years = sorted(medal_summary["पदक का वर्ष"].dropna().unique())
        selected_year = st.selectbox("📅 पदक वर्ष चुनें", options=["All"] + list(years))

        if selected_year != "All":
            medal_summary = medal_summary[medal_summary["पदक का वर्ष"] == selected_year]

        # Show table
        st.dataframe(
            medal_summary,
            use_container_width=True,
            hide_index=True

        )



